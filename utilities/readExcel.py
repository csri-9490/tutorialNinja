import openpyxl
wbook = openpyxl.load_workbook("C:/Users/Srikanth Chelukala/Desktop/ddriven.xlsx")
sheet = wbook.active
rows = sheet.max_row
columns = sheet.max_column
for row in range(1, rows+1):
    for column in range(1, columns):
        print(sheet.cell(row=row,column=column).value,end=" ")
    print()

for row in range(2, rows+1):
    for column in range(1, columns):
        print(sheet.cell(row=row,column=column).value)

final_list=[]
for r in range(2,rows+1):
        row_list=[]
        for c in range(1,columns+1):
            row_list.append(sheet.cell(row=r,column=c).value)
        final_list.append(row_list)
print('final list',final_list)

