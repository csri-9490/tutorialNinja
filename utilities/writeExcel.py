import openpyxl
wbook = openpyxl.load_workbook("C:/Users/Srikanth Chelukala/Desktop/ddriven.xlsx")
sheet = wbook.active
rows = sheet.max_row
columns = sheet.max_column

sheet.cell(row=1,column=3).value="Age"
for row in range(4,8):
    for col in range(1,4):
        sheet.cell(row=row,column=col).value=34
wbook.save("C:/Users/Srikanth Chelukala/Desktop/ddriven.xlsx")