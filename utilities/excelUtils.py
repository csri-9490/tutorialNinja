import openpyxl
import pytest

def get_row_count(path,sheet_name):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheet_name]
    return sheet.max_row
def get_column_count(path,sheet_name):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheet_name]
    return sheet.max_column

def get_cell_data(path,sheet_name,row_number,column_number):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheet_name]
    return sheet.cell(row=row_number,column=column_number).value

def set_cell_data(path,sheet_name,row_number,column_number,data):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheet_name]
    sheet.cell(row=row_number,column=column_number).value=data
    workbook.save(path)

def get_data_from_excel(path,sheet_name):
    final_list=[]
    workbook=openpyxl.load_workbook(path)
    sheet=workbook[sheet_name]
    total_rows=sheet.max_row
    total_cols=sheet.max_column

    for r in range(2,total_rows+1):
        row_list=[]
        for c in range(1,total_cols+1):
            row_list.append(sheet.cell(row=r,column=c).value)
        final_list.append(row_list)
    return final_list

# @pytest.fixture
def exceldata(path,sheet_name):
    u1=[]
    p1=[]
    wbook=openpyxl.load_workbook(path)
    sheet=wbook[sheet_name]#testdatas
    total_rows1 = sheet.max_row
    total_cols1 = sheet.max_column
    rows = get_row_count(path,sheet_name)
    for r in range(2,rows + 1):
       user=get_cell_data(path, sheet_name, r, 1)
       password =get_cell_data(path, sheet_name, r, 2)
       u1.append(user)
       p1.append(password)
    # print(u1)
    # print(p1)
    #path="C:\\Users\\Srikanth Chelukala\\PycharmProjects\\practNinja\\ExcelFiles\\tdata1.xlsx








