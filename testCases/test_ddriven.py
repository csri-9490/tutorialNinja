import logging

import openpyxl
import time

import pytest
from selenium.webdriver.common.by import By
from selenium import webdriver

from utilities.LogUtil import Logger
from utilities.generatingLogs import log, logger


log=Logger(__name__,logging.INFO)
@pytest.fixture(scope='module')
def setup():
    print("Creating  DB Concection")
    yield
    print("Closing DB Conncetion")

@pytest.fixture(scope='function')
def before():
    print("Launching browser")
    yield
    print("Closing Browser")
# @pytest.mark.parametrize("email,pwd",(['reddy.csri@gmail.com','Shivaya1@'],['reddy.csri1@gmail.com','Shivaya2@']))
# def get_data():
#     return [('reddy.csri@gmail.com','Shivaya1@'),('reddy.csri1@gmail.com','Shivaya2@')]
# @pytest.mark.parametrize("email,pwd",get_data())
# def test_dologin(setup,before,email,pwd):
#     driver = webdriver.Chrome()
#     driver.get("https://tutorialsninja.com/demo/index.php?route=account/login")
#     driver.find_element(By.ID, "input-email").send_keys(email)
#     driver.find_element(By.ID, "input-password").send_keys(pwd)
#     time.sleep(5)
#     driver.find_element(By.XPATH, "//input[@value='Login']").click()
#     driver.find_element(By.LINK_TEXT, "Logout").click()
    #//

# wbook = openpyxl.load_workbook("C:/Users/Srikanth Chelukala/Desktop/ddriven.xlsx")
# sheet = wbook.active
# rows = sheet.max_row
# columns = sheet.max_column
# for row in range(2, rows+1):
#     for column in range(1, columns):
#         print(sheet.cell(row=row,column=column).value)


def dta():
    wbook = openpyxl.load_workbook("C:/Users/Srikanth Chelukala/Desktop/ddriven.xlsx")
    sheet = wbook.active
    final_list=[]
    rows = sheet.max_row
    columns = sheet.max_column
    for r in range(2,rows+1):
        row_list=[]
        for c in range(1,columns+1):
            row_list.append(sheet.cell(row=r,column=c).value)
        final_list.append(row_list)
    return  final_list

@pytest.mark.parametrize("email,pwd",dta())
def test_dologin(setup,before,email,pwd):
    driver = webdriver.Chrome()
    log.logger.info("log info details,Test Started")
    driver.get("https://tutorialsninja.com/demo/index.php?route=account/login")
    driver.find_element(By.ID, "input-email").send_keys(email)
    driver.find_element(By.ID, "input-password").send_keys(pwd)
    time.sleep(5)
    driver.find_element(By.XPATH, "//input[@value='Login']").click()
    time.sleep(5)
    driver.find_element(By.LINK_TEXT, "Logout").click()
    log.logger.info("log info details,Test Ended")




