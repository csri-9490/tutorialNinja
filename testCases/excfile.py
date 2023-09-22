import openpyxl

from utilities.excelUtils import get_row_count, get_cell_data


def exceldata(path,sheet_name):
    u1=[]
    p1=[]
    wbook=openpyxl.load_workbook("C:\\Users\\Srikanth Chelukala\\PycharmProjects\\practNinja\\ExcelFiles\\tdata1.xlsx")
    sheet=wbook['testdatas']
    total_rows1 = sheet.max_row
    total_cols1 = sheet.max_column
    rows = get_row_count(path,sheet_name)
    for r in range(2,rows + 1):
       user=get_cell_data(path, sheet_name, r, 1)
       password =get_cell_data(path, sheet_name, r, 2)
       u1.append(user)
       p1.append(password)
    print(u1)
    print(p1)
path="C:\\Users\\Srikanth Chelukala\\PycharmProjects\\practNinja\\ExcelFiles\\tdata1.xlsx"
sheet_name="testdatas"
exceldata(path,sheet_name)
